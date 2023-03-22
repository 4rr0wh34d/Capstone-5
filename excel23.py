import os
import openpyxl
import sqlite3
# import tkinter as tk
from tkinter import messagebox
# import tabulate


class Excel23:

    def __init__(self):
        self.database_name = ''
        self.table_name = ''
        self.workbook_name = ''
        self.table_headings = ''
        self.db = ''
        self.cursor = ''
        # self.root_window = tk.Tk()

    def excel_to_sqlite(self, database_name, table_name, workbook_name):

        self.database_name = database_name
        self.table_name = table_name
        self.workbook_name = workbook_name

        db_name_str = './' + self.database_name
        if not os.path.exists(db_name_str):
            messagebox.showinfo('Database Error', 'Database does not exist')
        else:

            self.db = sqlite3.connect(self.database_name)
            self.cursor = self.db.cursor()

            # Assigning the variable pointing to the Excel workbook called 'books'
            wb = openpyxl.load_workbook(f'{self.workbook_name}')

            # Selecting the worksheets book_records. Alternatively can be done as ws = wb.active to select the first one
            # or the only one that exist
            # ws = wb[f'{worksheet_name}']
            ws = wb.active

            for i in range(2, ws.max_row + 1):
                values = [ws.cell(row=i, column=j).value for j in range(1, ws.max_column + 1)]
                print(values)
                self.cursor.execute(f'insert into {self.table_name} values(?, ?, ?, ?)', values)
            self.db.commit()
            self.db.close()

    def sqlite_to_excel(self, database_name, table_name, workbook_name, table_headings):
        self.database_name = database_name
        self.table_name = table_name
        self.workbook_name = workbook_name
        self.table_headings = table_headings
        
        if not os.path.exists('./' + self.database_name):
            messagebox.showinfo('Database Error', 'Database does not exist')
        else:

            try:
                self.db = sqlite3.connect(self.database_name)
                self.cursor = self.db.cursor()
    
                # Creating a new Excel workbook
                wb = openpyxl.Workbook()
                # wb.create_sheet('book_database')
                
                # Creating a handle to a new worksheet
                ws = wb.active  # wb['book_database']

                # writing the heading of the database
                for i, val in enumerate(self.table_headings):
                    ws.cell(row=1, column=i+1, value=val)
        
                self.cursor.execute(f'select * from  {self.table_name}')
        
                # Writing each row's cells to a worksheet with handle ws1
                for i, row in enumerate(self.cursor):
                    for j, col in enumerate(row):
        
                        ws.cell(row=i+2, column=j+1, value=col)
        
                # Saving the database
                wb.save(self.workbook_name)

            except Exception as e:
                messagebox.showinfo('Writing error', f'Error while writing to Excel file. {e}')
